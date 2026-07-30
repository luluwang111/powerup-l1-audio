#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""并发版: 根据【原版核对版】词汇表重新生成 PU1 配套音频 (Ryan 英式男声)。
- 读取: PowerUp_L1_词汇_原版核对版_最终版_音频录制.xlsx -> 全册词汇
- 列: 单元(0) 单词(3) 完整例句(9)
- 每个单元生成 正序.mp3 + 乱序.mp3，节奏: 单词→停1秒→例句→停2秒
- 乱序用固定种子(42)，保证可复现
- 并发生成: 同时 10 组 TTS 调用，避免逐词串行被网络往返拖慢
输出到脚本同目录 (pu1_web)。
"""
import asyncio, subprocess, os, random, re, shutil
from openpyxl import load_workbook
import edge_tts

XLSX = "/Users/zzymima0000/Desktop/PowerUp词汇抽取_原版核对/PowerUp_L1_词汇_原版核对版_最终版_音频录制.xlsx"
VOICE = "en-GB-RyanNeural"
INTRA = 1.0      # 单词 → 例句 停顿(秒)
ENTRY = 2.0      # 词条之间 停顿(秒)
SEED = 42
CONCURRENCY = 10  # 同时并发的 TTS 调用组数
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_BASE = "/tmp/pu1_seg_audio"  # 临时段文件放工作区外，避免触发批量删除保护
FFMPEG = shutil.which("ffmpeg") or "/opt/homebrew/bin/ffmpeg"


async def tts_save(text, path):
    await edge_tts.Communicate(text, VOICE).save(path)


def silence(path, secs):
    subprocess.run([FFMPEG, "-y", "-f", "lavfi", "-i", "anullsrc=r=24000:cl=mono",
                    "-t", str(secs), "-q:a", "9", "-acodec", "libmp3lame", path],
                   capture_output=True)


async def build_unit(unit, items, out_path, base_tmp, sem):
    safe = re.sub(r"[^A-Za-z0-9_-]", "_", str(unit))
    tmp = os.path.join(base_tmp, safe)
    os.makedirs(tmp, exist_ok=True)
    wfiles = [None] * len(items)
    sfiles = [None] * len(items)

    async def gen_one(i, w, s):
        async with sem:
            wp = os.path.join(tmp, f"w{i}.mp3")
            sp = os.path.join(tmp, f"s{i}.mp3")
            await tts_save(w, wp)
            await tts_save(s, sp)
        wfiles[i] = wp
        sfiles[i] = sp

    await asyncio.gather(*[gen_one(i, w, s) for i, (w, s) in enumerate(items)])

    listf = os.path.join(tmp, "list.txt")
    with open(listf, "w") as f:
        for idx in range(len(items)):
            f.write(f"file '{wfiles[idx]}'\n")
            f.write(f"file '{os.path.join(base_tmp, 'intra.mp3')}'\n")
            f.write(f"file '{sfiles[idx]}'\n")
            if idx != len(items) - 1:
                f.write(f"file '{os.path.join(base_tmp, 'entry.mp3')}'\n")
    subprocess.run([FFMPEG, "-y", "-f", "concat", "-safe", "0", "-i", listf,
                    "-c", "copy", out_path], capture_output=True)
    print(f"  ✅ {unit}: {len(items)} 词 已拼接")


async def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["全册词汇"]
    rows = list(ws.iter_rows(values_only=True))
    order = []
    groups = {}
    skipped = 0
    for r in rows[1:]:
        unit = r[0]
        word = r[3]
        sentence = r[9]
        if not unit or not word or not sentence:
            skipped += 1
            continue
        unit = str(unit).strip()
        if unit not in groups:
            groups[unit] = []
            order.append(unit)
        groups[unit].append((str(word).strip(), str(sentence).strip()))

    os.makedirs(TMP_BASE, exist_ok=True)
    silence(os.path.join(TMP_BASE, "intra.mp3"), INTRA)
    silence(os.path.join(TMP_BASE, "entry.mp3"), ENTRY)
    sem = asyncio.Semaphore(CONCURRENCY)

    print(f"共 {len(order)} 个单元段: {order}  (跳过空行 {skipped})")
    for unit in order:
        items = groups[unit]
        await build_unit(unit, items, os.path.join(OUT_DIR, f"{unit}_正序.mp3"), TMP_BASE, sem)
        shuf = items[:]
        random.Random(SEED).shuffle(shuf)
        await build_unit(unit, shuf, os.path.join(OUT_DIR, f"{unit}_乱序.mp3"), TMP_BASE, sem)
        print(f"  🎉 {unit} 完成 (正序+乱序)")
    print("全部生成完成 (临时段文件留在 /tmp/pu1_seg_audio)")


if __name__ == "__main__":
    asyncio.run(main())
