#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据【原版核对版】词汇表重新生成 PU1 配套音频 (Ryan 英式男声)。
- 读取: PowerUp_L1_词汇_原版核对版_最终版_音频录制.xlsx -> 全册词汇
- 列: 单元(0) 单词(3) 完整例句(9)
- 每个单元生成 正序.mp3 + 乱序.mp3，节奏: 单词→停1秒→例句→停2秒
- 乱序用固定种子(42)，保证可复现
输出到脚本同目录 (pu1_web)。
"""
import asyncio, subprocess, os, random, shutil
from openpyxl import load_workbook
import edge_tts

XLSX = "/Users/zzymima0000/Desktop/PowerUp词汇抽取_原版核对/PowerUp_L1_词汇_原版核对版_最终版_音频录制.xlsx"
VOICE = "en-GB-RyanNeural"
INTRA = 1.0   # 单词 → 例句 停顿(秒)
ENTRY = 2.0   # 词条之间 停顿(秒)
SEED = 42
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_DIR = "/tmp/pu1_seg_audio"  # 临时段文件放工作区外，避免触发批量删除保护
FFMPEG = shutil.which("ffmpeg") or "/opt/homebrew/bin/ffmpeg"


async def tts_save(text, path):
    await edge_tts.Communicate(text, VOICE).save(path)


def silence(path, secs):
    subprocess.run([FFMPEG, "-y", "-f", "lavfi", "-i", "anullsrc=r=24000:cl=mono",
                    "-t", str(secs), "-q:a", "9", "-acodec", "libmp3lame", path],
                   capture_output=True)


async def build_unit(items, out_path, tmp):
    wfiles = []
    sfiles = []
    for i, (w, s) in enumerate(items):
        wp = os.path.join(tmp, f"w{i}.mp3")
        sp = os.path.join(tmp, f"s{i}.mp3")
        await tts_save(w, wp)
        await tts_save(s, sp)
        wfiles.append(wp)
        sfiles.append(sp)
    listf = os.path.join(tmp, "list.txt")
    with open(listf, "w") as f:
        for idx in range(len(items)):
            f.write(f"file '{wfiles[idx]}'\n")
            f.write(f"file '{os.path.join(tmp, 'intra.mp3')}'\n")
            f.write(f"file '{sfiles[idx]}'\n")
            if idx != len(items) - 1:
                f.write(f"file '{os.path.join(tmp, 'entry.mp3')}'\n")
    subprocess.run([FFMPEG, "-y", "-f", "concat", "-safe", "0", "-i", listf,
                    "-c", "copy", out_path], capture_output=True)
    # 临时段文件留在 TMP_DIR (/tmp)，由系统自行清理，避免触发批量删除保护


async def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["全册词汇"]
    rows = list(ws.iter_rows(values_only=True))
    order = []
    groups = {}
    skipped = 0
    for r in rows[1:]:
        unit = r[0]
        word = r[3]
        sentence = r[9]
        if not unit or not word or not sentence:
            skipped += 1
            continue
        unit = str(unit).strip()
        if unit not in groups:
            groups[unit] = []
            order.append(unit)
        groups[unit].append((str(word).strip(), str(sentence).strip()))

    tmp = TMP_DIR
    os.makedirs(tmp, exist_ok=True)
    silence(os.path.join(tmp, "intra.mp3"), INTRA)
    silence(os.path.join(tmp, "entry.mp3"), ENTRY)

    print(f"共 {len(order)} 个单元段: {order}  (跳过空行 {skipped})")
    for unit in order:
        items = groups[unit]
        await build_unit(items, os.path.join(OUT_DIR, f"{unit}_正序.mp3"), tmp)
        shuf = items[:]
        random.Random(SEED).shuffle(shuf)
        await build_unit(shuf, os.path.join(OUT_DIR, f"{unit}_乱序.mp3"), tmp)
        print(f"  ✅ {unit}: {len(items)} 词 (正序+乱序)")
    print("全部生成完成 (临时段文件留在 /tmp/pu1_seg_audio)")


if __name__ == "__main__":
    asyncio.run(main())
